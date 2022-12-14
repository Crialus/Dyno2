from pyexpat import model
from statistics import pstdev
import openpyxl as xl
import pandas as pd
import inquirer
from inquirer.themes import GreenPassion
import os
import re
import sys


def get_shd():
    # User input for SHD number that ensures correct format via regex
    # and return SHD
    shd = input('Please enter an SHD number in the form YYMMDDXX####: ')
    rex = re.compile('^[0-9]{6}[A-Z]{2}[0-9]{4}$')
    if rex.match(shd):
        return shd
    else:
        sys.exit("Incorrect SHD Format")


def get_unit_type():
    # Get folder containing results data. Load all folders from root and use
    # inquirer list to select the folder containing SHD and return the path
    # to the relevant SHD data
    ROOT = 'T:\\Motors\\Lab Testing\\00 Active Tasks\\'
    dir_list = [item for item in os.listdir(ROOT) if os.path.isdir(os.path.join(ROOT, item))]
    model_list = [
        inquirer.List('model',
                      message="What type of unit is it? ",
                      choices=dir_list,
                      carousel=True),
        ]
    model = inquirer.prompt(model_list, theme=GreenPassion()).get('model')
    path = ROOT + model
    return (path)


def results_files(path, shd):
    # Create paths for all required files and return a dictionary of them
    template_path = path + '\\' + 'Template'
    path = path + '\\' + shd
    unit_paths = {
        'template': template_path + '\\' + 'Template.xlsx',
        'results': path + '\\' + shd + '.xlsx',
        'cogging': path + '\\Cogging_LSF\\' + shd + '_50 - Friction Results.dsv',
        'cogging_fix_only': path + '\\Cogging_LSF\\' + shd + '_FixOnly_50 - Friction Results.dsv',
        'high_speed_friction': path + '\\HSF\\' + shd + '_1000 - Friction Results.dsv',
        'high_speed_friction_fix_only': path + '\\HSF\\'+shd+'_FixOnly_1000 - Friction Results.dsv',
        'back_emf_123': path + '\\BEMF Ke\\' + shd + '_123 Results.csv',
        'back_emf_456': path + '\\BEMF Ke\\' + shd + '_456 Results.csv',
        'back_emf_789': path + '\\BEMF Ke\\' + shd + '_789 Results.csv',
        'back_emf_101112': path + '\\BEMF Ke\\' + shd + '_101112 Results.csv',
        'mps': path + '\\MPS_Flux\\' + shd + ' Results.csv',
        }
    return unit_paths


class Transfer:
    def __init__(self, paths, shd):
        # If results spreadsheet exists then use that as template base, else
        # load a new sheet from template
        self.paths = paths
        self.shd = shd
        self.template = self.paths.get('template')
        self.results = self.paths.get('results')
        if os.path.exists(self.results):
            self.entry = xl.load_workbook(self.results)
        else:
            self.entry = xl.load_workbook(self.template)

    def misc(self):
        # Add the SHD to the serial number of generated report and label
        # the rows in cogging for both unit and fixture only
        sheet = self.entry['Generated Report']
        sheet.cell(row=10, column=4, value=self.shd)
        sheet = self.entry['Cogging']
        sheet.cell(row=10, column=10, value=self.shd)
        sheet.cell(row=11, column=10, value=(self.shd+'_fix_only'))
        self.entry.save(self.results)

    def copy_data(self, sheet, static, variable, type, values):
        # copy data held in variables to the results file, for multiple
        # values either row or column will be variable and the
        # passed 'type' will define which this is.
        # the passed 'variable' is the start number for the iteration
        i = variable
        sheet = self.entry[sheet]
        if type == 'row':
            for x in values:
                x = float(x)
                sheet.cell(row=i, column=static, value=x)
                i += 1
        elif type == 'column':
            for x in values:
                x = float(x)
                sheet.cell(row=static, column=i, value=x)
                i += 1
        else:
            x = values
            sheet.cell(row=static, column=variable, value=x)

    def cogging(self):
        # For fixture only, take the inline mean values and add
        # them to the results. For the unit data take inline
        # mean values, and peak to peak values
        sheet = 'Cogging'
        fix_only = self.paths.get('cogging_fix_only')
        cogging = self.paths.get('cogging')
        fix_only_data = pd.DataFrame(pd.read_csv(open(fix_only), delimiter='\t', header=None))
        fix_only_inline_mean = (fix_only_data.loc[15, 1], fix_only_data.loc[15, 4])
        self.copy_data(sheet, 11, 22, 'column', fix_only_inline_mean)
        cogging_data = pd.DataFrame(pd.read_csv(open(cogging), delimiter='\t', header=None))
        cogging_p2p = (cogging_data.loc[14, 1], cogging_data.loc[14, 4])
        cogging_inline_mean = (cogging_data.loc[15, 1], cogging_data.loc[15, 4])
        cogging_cw = cogging_data.loc[23:70, 1]
        cogging_acw = cogging_data.loc[23:70, 4]
        # test_date = str(cogging_data.loc[2, 2])
        # self.copy_data(sheet, 2, 2, 'single', test_date)
        self.copy_data(sheet, 10, 22, 'column', cogging_inline_mean)
        self.copy_data(sheet, 10, 28, 'column', cogging_p2p)
        self.copy_data(sheet, 10, 32, 'column', cogging_cw)
        self.copy_data(sheet, 10, 105, 'column', cogging_acw)
        self.entry.save(self.results)

    def hsf(self):
        # Take inline means for both unit and fixture only data at 1000 rpm
        sheet = 'High Speed Friction'
        fix_only = self.paths.get('high_speed_friction_fix_only')
        hsf = self.paths.get('high_speed_friction')
        fix_only_data = pd.DataFrame(pd.read_csv(open(fix_only), delimiter='\t', header=None))
        fix_only_inline_mean = (fix_only_data.loc[15, 1], fix_only_data.loc[15, 4])
        self.copy_data(sheet, 20, 2, 'column', fix_only_inline_mean)
        hsf_data = pd.DataFrame(pd.read_csv(open(hsf), delimiter='\t', header=None))
        hsf_inline_mean = (hsf_data.loc[15, 1], hsf_data.loc[15, 4])
        self.copy_data(sheet, 8, 2, 'column', hsf_inline_mean)
        self.entry.save(self.results)

    def bemf(self):
        # Take mean Ke and Ke betweeb each phase, also take
        # rotational speed during test
        sheet = 'Bemf Ke'
        bemf123 = self.paths.get('back_emf_123')
        bemf456 = self.paths.get('back_emf_456')
        bemf789 = self.paths.get('back_emf_789')
        bemf101112 = self.paths.get('back_emf_101112')
        bemf123_data = pd.DataFrame(pd.read_csv(open(bemf123), delimiter=',', header=None))
        bemf123_cw_values = bemf123_data.loc[12:16, 1]
        bemf123_acw_values = bemf123_data.loc[12:16, 4]
        self.copy_data(sheet, 3, 15, 'row', bemf123_cw_values)
        self.copy_data(sheet, 6, 15, 'row', bemf123_acw_values)
        bemf456_data = pd.DataFrame(pd.read_csv(open(bemf456), delimiter=',', header=None))
        bemf456_cw_values = bemf456_data.loc[12:16, 1]
        bemf456_acw_values = bemf456_data.loc[12:16, 4]
        self.copy_data(sheet, 10, 15, 'row', bemf456_cw_values)
        self.copy_data(sheet, 13, 15, 'row', bemf456_acw_values)
        # If motor has lanes 3 + 4 then add the same values to the sheet.
        # Needs an additional solution as currently data is simply stored
        # in the same sheet and no analysis takes place, potential to
        # either create new file and store in the analysed cells or needs
        # new template
        if os.path.exists(bemf789):
            bemf789_data = pd.DataFrame(pd.read_csv(open(bemf789), delimiter=',', header=None))
            bemf789_cw_values = bemf789_data.loc[12:16, 1]
            bemf789_acw_values = bemf789_data.loc[12:16, 4]
            self.copy_data(sheet, 4, 15, 'row', bemf789_cw_values)
            self.copy_data(sheet, 7, 15, 'row', bemf789_acw_values)
        if os.path.exists(bemf101112):
            bemf101112_data = pd.DataFrame(pd.read_csv(open(bemf101112), delimiter=',', header=None))
            bemf101112_cw_values = bemf101112_data.loc[12:16, 1]
            bemf101112_acw_values = bemf101112_data.loc[12:16, 4]
            self.copy_data(sheet, 11, 15, 'row', bemf101112_cw_values)
            self.copy_data(sheet, 14, 15, 'row', bemf101112_acw_values)
        self.entry.save(self.results)

    def mps_dipole(self):
        sheet = 'MPS'
        mps = self.paths.get('mps')
        mps_data = pd.DataFrame(pd.read_csv(open(mps), delimiter=',', header=None))
        mps_cw_field = mps_data.loc[12:14, 1]
        mps_acw_field = mps_data.loc[12:14, 4]
        mps_cw_error = mps_data.loc[16:19, 1]
        mps_acw_error = mps_data.loc[16:19, 4]
        mps_cw_flux_x = mps_data.loc[22:23, 1]
        mps_cw_flux_y = mps_data.loc[24:25, 1]
        mps_acw_flux_x = mps_data.loc[22:23, 4]
        mps_acw_flux_y = mps_data.loc[24:25, 4]
        self.copy_data(sheet, 12, 3, 'column', mps_cw_flux_x)
        self.copy_data(sheet, 13, 3, 'column', mps_acw_flux_x)
        self.copy_data(sheet, 12, 6, 'column', mps_cw_flux_y)
        self.copy_data(sheet, 13, 6, 'column', mps_acw_flux_y)
        self.copy_data(sheet, 20, 3, 'column', mps_cw_error)
        self.copy_data(sheet, 21, 3, 'column', mps_acw_error)
        self.copy_data(sheet, 20, 6, 'column', mps_cw_field)
        self.copy_data(sheet, 21, 6, 'column', mps_acw_field)
        self.entry.save(self.results)


def main():
    shd = get_shd()
    path = get_unit_type()
    unit_paths = results_files(path, shd)
    unit = Transfer(unit_paths, shd)
    unit.cogging()
    unit.hsf()
    unit.bemf()
    unit.misc()
    unit.mps_dipole()


main()
